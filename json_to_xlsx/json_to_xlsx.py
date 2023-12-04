import json
import openpyxl

json_folder_path = "/home/zsy/code/json_to_xlsx/"           # 设置需要读取的JSON文件夹路径
json_file_name = "robot_fault_Zh.json"                          # 读取的JSON文件名称
json_file_absolute_path = json_folder_path + json_file_name     # JSON文件绝对路径

xlsx_folder_path = json_folder_path                             # 生成的xlsx文件保存路径
xlsx_file_name = "json_to_xlsx.xlsx"                            # 读取的xlsx文件名称
xlsx_file_absolute_path = xlsx_folder_path + xlsx_file_name     # xlsx文件绝对路径

print(json_file_name)
with open(json_file_absolute_path, "r") as f:
    data = json.load(f)
    
    # 获取第四个键值对之后的所有数据
    data = dict(list(data.items())[4:])
    keys = data.keys()
    key_list = list(keys)       
    print(f"\n获取到的key_list：'{key_list}'\n")

# 创建工作簿
workbook = openpyxl.Workbook()

# 遍历数据的键
for key in data.keys():
    print(f"正在转换 : '{key}'")

    sheet_name = key[:31]  # 限制工作表名称的最大长度为31个字符
    worksheet = workbook.create_sheet(title=sheet_name)  # 创建工作表

    # 保存description以及fault_type,至xlsx的第一行与第二行
    worksheet.cell(row=1, column=1).value = "description"
    worksheet.cell(row=1, column=2).value = "fault_type"
    worksheet.cell(row=2, column=1).value = data[key]["description"]
    worksheet.cell(row=2, column=2).value = data[key]["fault_type"]

    # 循环读取fault_info中的数据，并保存在xlsx文件中
    fault_info = data[key].get("fault_info")

    if not fault_info:
        print(f"           '{key}' 中的 'fault_info' 为空，跳过相关操作。")
        continue

    # 写入表头
    headers = list(fault_info[0].keys())
    for col_num, header in enumerate(headers, 1):       
        worksheet.cell(row=1+2, column=col_num).value = header

    # 写入数据
    for row_num, fault_data in enumerate(fault_info, 2+2):
        for col_num, key in enumerate(headers, 1):
            worksheet.cell(row=row_num, column=col_num).value = fault_data[key]

# 删除默认创建的工作表
default_sheet = workbook["Sheet"]
workbook.remove(default_sheet)


# 保存文件
workbook.save(xlsx_file_absolute_path)
print(f"\n文件已保存为: {xlsx_file_absolute_path}")
