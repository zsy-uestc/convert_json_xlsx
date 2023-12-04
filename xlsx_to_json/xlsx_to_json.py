import openpyxl
import json

xlsx_folder_path = "/home/zsy/code/json_to_xlsx/"               # 需要转换的xlsx文件夹路径
xlsx_file_name = "json_to_xlsx.xlsx"                            # 读取的xlsx文件名称
xlsx_file_absolute_path = xlsx_folder_path + xlsx_file_name     # xlsx文件绝对路径

json_folder_path = xlsx_folder_path                             # 生成的json文件保存路径
json_file_name = "xlsx_to_json.json"                            # 读取的json文件名称
json_file_absolute_path = json_folder_path + json_file_name     # json文件绝对路径

def xlsx_to_json(filename):
    workbook = openpyxl.load_workbook(filename)
    sheets = workbook.sheetnames

    # 转换后的内容，固定的显示内容
    data = {
        "文档描述1": "故障说明:xxxxxxx",
        "文档描述2": "各自模块根据故障分类添加各自的故障项及故障描述信息及故障处理方式",
        "description": {
            "err_code": "错误码共8位,从10000000-FFFFFFFF,xxxxxx"
        },
        "fault_version": "xxxxx"
    }

    for sheet_name in sheets:

        sheet = workbook[sheet_name]
        key1 = [cell.value for cell in sheet[1]]        # 读取xlsx文件中第一行前两列的内容，为key                   --> ['description', 'fault_type']
        key1 = key1[0:2]
        value1 = [cell.value for cell in sheet[2]]      # 读取xlsx文件中第二行前两列的内容，为第一行key所对应的value  --> ['control 节点错误信息', '01']
        value1 = value1[0:2]   

        dict = {}
        for key, val in zip(key1, value1):
            dict[key] = val

        keys = [cell.value for cell in sheet[3]]        # 读取xlsx文件中第三行所有列的内容，为key         --> ['fault_devel', 'code', 'descr', 'solution']
        values = []
        for row in range(4, sheet.max_row + 1):         # 从第四行开始遍历每一行, 为第三行key所对应的value --> ['1', '10100000', 'task start', '开始任务']
            value = [cell.value for cell in sheet[row]]
            values.append(value)
        items = []
        for value in values:
            item = {}
            for key, val in zip(keys, value):
                if val is None or value == "null":      # 如果读取到的是null，显示为""
                    val = ""

                item[key] = val
            items.append(item)
        data[sheet_name] = dict
        dict["fault_info"] = items
    json_data = json.dumps(data, indent=4, ensure_ascii=False)  
    with open(json_file_absolute_path, 'w', encoding='utf-8') as f:
        f.write(json_data)
    print(f"\n文件已保存为: {xlsx_file_absolute_path}" )

# 调用函数，传入Excel文件路径
xlsx_to_json(xlsx_file_absolute_path)
